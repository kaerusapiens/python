serach_mall_value = input("Enter value code : ") 
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoSuchAttributeException
from random_user_agent.user_agent import UserAgent
from random_user_agent.params import SoftwareName, OperatingSystem
from selenium.webdriver.common.proxy import Proxy, ProxyType
from selenium.webdriver.common.action_chains import ActionChains
import sys
import time
import random
import openpyxl
from openpyxl import Workbook
import os
import os.path

def html_save_to_path(folder_name, file_name, html_file):
    folder_name=folder_name.replace("/","")
    file_name=file_name.replace("/","")
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
        print(f"Folder '{folder_name}' created.")
        file_path = os.path.join(folder_name, file_name)
        with open(file_path, mode='w', encoding = 'utf-8') as file:
            file.write(html_file)
    else:
        #print(f"Folder '{folder_name}' already exists.")
        file_path = os.path.join(folder_name, file_name)
        with open(file_path, mode='w', encoding = 'utf-8') as file:
            file.write(html_file)
   
random_wait_time = random.uniform(5, 10)
software_names = [SoftwareName.CHROME.value]
operating_systems = [OperatingSystem.WINDOWS.value, OperatingSystem.LINUX.value]   
user_agent_rotator = UserAgent(software_names=software_names, operating_systems=operating_systems, limit=100)
user_agent = user_agent_rotator.get_random_user_agent()
print(user_agent)
proxy_ip ="XXXX"
proxy_port = "XXXXXXXX"
proxy_username = "XXXXXXXXXXXXXXXX"
proxy_pw = "XXXXXXXXXXXXXx"
proxy = Proxy()
proxy.proxy_type = ProxyType.MANUAL
proxy.http_proxy = f"{proxy_ip}:{proxy_port}"
proxy.ssl_proxy = f"{proxy_ip}:{proxy_port}"
proxy.socks_proxy = f"{proxy_ip}:{proxy_port}"
proxy.socks_username = proxy_username
proxy.socks_password = proxy_pw
options = Options()
options.add_argument('lang=ko_KR')
options.add_argument("--start-maximized")
options.add_argument("--no-sandbox")
options.add_argument('--ignore-certificate-errors')
options.add_argument("user-agent="+user_agent)
options.add_argument("accept=text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8")
options.add_argument("accept-charset=cp1254,ISO-8859-9,utf-8;q=0.7,*;q=0.3")
options.add_argument("accept-encoding=gzip,deflate,sdch")
options.add_argument("accept-language=tr,tr-TR,en-US,en;q=0.8")
options.add_experimental_option('useAutomationExtension', False)
options.add_argument("--start-maximized")
options.add_argument("--proxyserver=http://{}:{}@{}:{}".format(proxy_username,proxy_pw,proxy_ip,proxy_port))
print("--proxyserver=http://{}:{}@{}:{}".format(proxy_username,proxy_pw,proxy_ip,proxy_port))
#options.add_argument('--incognito')
options.add_experimental_option("detach", True)
options.add_argument("--headless=new") 

# Create a Proxy object
proxy = Proxy({
    'proxyType': ProxyType.MANUAL,
    'httpProxy': f"{proxy_host}:{proxy_port}",
    'ftpProxy': f"{proxy_host}:{proxy_port}",
    'sslProxy': f"{proxy_host}:{proxy_port}",
    'noProxy': ''
})

# Add proxy authentication if needed
if proxy_username and proxy_password:
    proxy.proxy_type = ProxyType.MANUAL
    proxy.http_proxy = f"{proxy_username}:{proxy_password}@{proxy_host}:{proxy_port}"
driver = webdriver.Chrome(options=options)  
driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument",{ "source": """ Object.defineProperty(navigator, 'webdriver', { get: () => undefined }) """ })  
#driver = webdriver.Remote(command_executor='http://selenium:4444/wd/hub',options=options)
#driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument",{ "source": """ Object.defineProperty(navigator, 'webdriver', { get: () => undefined }) """ })
driver.get("https://www.coupang.com/")
time.sleep(random_wait_time)
wait = WebDriverWait(driver, 2)    
MAX_WAIT_TIME=4
print(driver.title)



wbb = openpyxl.load_workbook('all_mall.xlsx')
wss = wbb.active


for row in wss.iter_rows(min_row=2): # type: 
    wb = Workbook()
    ws = wb.active
    ws.append(["mall_code","mall_name","num","cate1_code","cate1_name","num_2","cate2_code","cate2_name","cate3_code","cate3_name","cate4_code","cate4_name"]) # type: ignore
    if str(row[1].value) == str(serach_mall_value): 
        print("start... :" , row[2].value)
        #sys.stdout = open("output_"+str(row[1].value)+"_"+row[0].value.replace("/","")+'.txt','wt')
        try:
            #자동차용품-------------------------------Inputput mall & donwload cate1-------------------------------------------------
            #1.go to menu bar
            print("menu bar selected...")
            element_to_wait_for = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#header > div')))
            header_click= driver.find_element(By.CSS_SELECTOR,'#header > div')
            actions = ActionChains(driver)
            actions.move_to_element(header_click)
            actions.perform()
            #driver.execute_script("arguments[0].click();", header_click)
            #element_to_wait_for.click()
            print("....completed")
            #2.click the mall cate name
            print("clicked mall cate name:",row[2].value)
            time.sleep(random_wait_time)
            driver.implicitly_wait(MAX_WAIT_TIME)
            element_to_wait_for = wait.until(EC.presence_of_element_located((By.LINK_TEXT, row[2].value)))
            mall_click= driver.find_element(By.LINK_TEXT, row[2].value)
            actions = ActionChains(driver)
            actions.move_to_element(mall_click)
            actions.perform()
            driver.execute_script("arguments[0].click();", mall_click)
            #element_to_wait_for.click()
            print("....completed")
            time.sleep(random_wait_time)
            driver.implicitly_wait(MAX_WAIT_TIME)
            #3.download source & save by category
            print("downloaded mall name:",row[2].value)
            html_content = driver.page_source
            html_save_to_path(str(row[1].value) + "_" + row[0].value, str(row[1].value) + "_" + row[0].value +'.html',html_content) # type: ignore
            print("....completed")
            #-------------------------------Get cate1---------------------------------------------------------------   
            index = 1  # Start with the first child
            while True:
                try: #겨울철 차량관리.try to get element&attirbute / except element& attribute anymore
                    time.sleep(random_wait_time)
                    driver.implicitly_wait(MAX_WAIT_TIME)
                    parent_selector = "#searchCategoryComponent > ul > li" #searchCategoryComponent > ul > li:nth-child(1)
                    #cate1code  #searchCategoryComponent > ul > li:nth-child(1)
                    cate1_selector_base = f"{parent_selector}:nth-child({index})"
                    
                    cate1_selector_click = driver.find_element(By.CSS_SELECTOR, cate1_selector_base)
                    print(cate1_selector_click )
                    cate1_data_component_id = cate1_selector_click.get_attribute("data-component-id")       
                    print(cate1_data_component_id )         
                    #cate1name #searchCategoryComponent > ul > li:nth-child(1) > label
                    cate1_label = f"{cate1_selector_base} > label"
                    cate1_label_element = driver.find_element(By.CSS_SELECTOR,cate1_label)     
                    print(cate1_label_element)            
                    print("*GET : ",index,"/",cate1_data_component_id,"/",cate1_label_element.text,"....completed")
                    ws.append([row[2].value, row[1].value, index, cate1_data_component_id, cate1_label_element.text, "", "", "","","","","","",""])
                    wb.save("list_"+str(row[1].value)+"_"+row[0].value.replace("/","")+".xlsx")
                    index += 1
                    #-------★click 열림
                    try:
                        #겨울철 차량관리.try to click 겨울철 차량관리 tab / timeout - need to go next tab. / no such element- loop end
                        # Check if the nth-child element contains an 'a' element
                        ##searchCategoryComponent > ul > li:nth-child(2) > a
                        print("trying to click 열림 button and go to cat1 ->cat2...")
                        time.sleep(random_wait_time)
                        driver.implicitly_wait(MAX_WAIT_TIME)
                        print(f"{cate1_selector_base} > a")
                        element_to_wait_for =  WebDriverWait(driver, MAX_WAIT_TIME).until(EC.presence_of_element_located((By.CSS_SELECTOR, f"{cate1_selector_base} > a")))
                        move_to_ele_1 = driver.find_element(By.CSS_SELECTOR, f"{cate1_selector_base} > a")
                        actions = ActionChains(driver)
                        actions.move_to_element(move_to_ele_1)
                        actions.perform()
                        driver.execute_script("arguments[0].click();", move_to_ele_1)
                        wb.save("list_"+str(row[1].value)+"_"+row[0].value.replace("/","")+".xlsx")
                        print("....click successufl! Go into Cate2")
                        #-------------------------------Get cate2---------------------------------------------------------------  
                        index_2 = 1
                        while True:
                            try: #핸들커버.try to get element&attirbute / except element& attribute anymore
                                time.sleep(random_wait_time)
                                driver.implicitly_wait(MAX_WAIT_TIME)
                                #1.download source & save by category
                                html_content = driver.page_source
                                html_save_to_path(str(row[1].value) + "_" + row[0].value, str(row[1].value)+"_"+str(cate1_data_component_id)+'.html',html_content)
                                #2.Get Gate2 # category176430 > li:nth-child(1) 
                                cate2_selector_base = f"#category{cate1_data_component_id} > li:nth-child({index_2})"
                                cate2_selector_click = driver.find_element(By.CSS_SELECTOR, cate2_selector_base)
                                cate2_data_component_id = cate2_selector_click.get_attribute('data-component-id')
                                #catename
                                cate2_label = f"{cate2_selector_base} > label"
                                cate2_label_element = driver.find_element(By.CSS_SELECTOR,cate2_label)
                                print("**GET : ", index_2,"/",cate2_data_component_id,"/",cate2_label_element.text,"....completed")     
                                ws.append([row[2].value, row[1].value,index, cate1_data_component_id, cate1_label_element.text, index_2, cate2_data_component_id, cate2_label_element.text,"", "", "","","",""])  
                                index_2 += 1
                                #-------★click 열림
                                try:
                                     #핸들커버..try to click핸들커버 tab / timeout - need to go next tab. / no such element- loop end
                                    # Check if the nth-child element contains an 'a' element
                                    ##category409204 > li:nth-child(2) > a #searchCategoryComponent > ul > li:nth-child(3) > a
                                    time.sleep(random_wait_time)
                                    driver.implicitly_wait(MAX_WAIT_TIME)
                                    print("trying to click 열림 button and go to cat2 ->cat3...")
                                    element_to_wait_for =  WebDriverWait(driver, MAX_WAIT_TIME).until(EC.presence_of_element_located((By.CSS_SELECTOR, f"{cate2_selector_base} > a")))
                                    move_to_ele_2 = driver.find_element(By.CSS_SELECTOR, f"{cate2_selector_base} > a")                                   
                                    actions = ActionChains(driver)
                                    actions.move_to_element(move_to_ele_2)
                                    actions.perform()
                                    driver.execute_script("arguments[0].click();", move_to_ele_2)
                                    print("....click successufl! Go into Cate3")
                                    #-------------------------------Get cate3---------------------------------------------------------------  
                                    index_3 = 1
                                    while True:
                                        try:
                                            time.sleep(random_wait_time)
                                            driver.implicitly_wait(MAX_WAIT_TIME)
                                            #1.download source & save by category
                                            html_content = driver.page_source
                                            html_save_to_path(str(row[1].value) + "_" + row[0].value, str(row[1].value)+"_"+str(cate1_data_component_id)+"_"+str(cate2_data_component_id)+'.html',html_content)
                                            #2.#category435460 > li:nth-child(1)
                                            cate3_selector_base = f"#category{cate2_data_component_id} > li:nth-child({index_3})"
                                            cate3_selector_click = driver.find_element(By.CSS_SELECTOR, cate3_selector_base)
                                            cate3_data_component_id = cate3_selector_click.get_attribute('data-component-id')
                                            #catename #category435460 > li:nth-child(1) > label
                                            cate3_label = f"{cate3_selector_base} > label"
                                            cate3_label_element = driver.find_element(By.CSS_SELECTOR,cate3_label)
                                            ws.append([row[2].value, row[1].value, index, cate1_data_component_id, cate1_label_element.text, index_2, cate2_data_component_id, cate2_label_element.text,index_3,cate3_data_component_id,cate3_label_element.text,"","",""])
                                            print("***GET : ", index_3,"/",cate3_data_component_id,"/",cate3_label_element.text,"....completed")
                                            index_3 += 1
                                            #-------★click 열림
                                            try:
                                                #시트.try to click시트 tab / timeout - need to go next tab. / no such element- loop end
                                                # Check if the nth-child element contains an 'a' element
                                                ##category409204 > li:nth-child(2) > a #searchCategoryComponent > ul > li:nth-child(3) > a
                                                time.sleep(2)
                                                driver.implicitly_wait(MAX_WAIT_TIME)
                                                print("trying to click 열림 button and go to cat3 ->cat4...")
                                                element_to_wait_for =  WebDriverWait(driver, MAX_WAIT_TIME).until(EC.presence_of_element_located((By.CSS_SELECTOR, f"{cate3_selector_base} > a")))
                                                move_to_ele_3 = driver.find_element(By.CSS_SELECTOR, f"{cate3_selector_base} > a")                                   
                                                actions = ActionChains(driver)
                                                actions.move_to_element(move_to_ele_3)
                                                actions.perform()
                                                driver.execute_script("arguments[0].click();", move_to_ele_3)
                                                print("....click successufl! Go into Cate3")
                                                #-------------------------------Get cate4---------------------------------------------------------------  
                                                index_4 = 1
                                                while True:
                                                    try:
                                                        time.sleep(random_wait_time)
                                                        driver.implicitly_wait(MAX_WAIT_TIME)
                                                        #1.download source & save by category
                                                        html_content = driver.page_source
                                                        html_save_to_path(str(row[1].value) + "_" + row[0].value, str(row[1].value)+"_"+str(cate1_data_component_id)+"_"+str(cate2_data_component_id)+"_"+str(cate3_data_component_id)+'.html',html_content)
                                                        #2.#category400935 > li:nth-child(1)
                                                        cate4_selector_base = f"#category{cate3_data_component_id} > li:nth-child({index_4})"
                                                        cate4_selector_click = driver.find_element(By.CSS_SELECTOR, cate4_selector_base)
                                                        cate4_data_component_id = cate4_selector_click.get_attribute('data-component-id')
                                                        #catename #category435460 > li:nth-child(1) > label
                                                        cate4_label = f"{cate4_selector_base} > label"
                                                        cate4_label_element = driver.find_element(By.CSS_SELECTOR,cate4_label)
                                                        ws.append([row[2].value, row[1].value, index, cate1_data_component_id, cate1_label_element.text, index_2, cate2_data_component_id, cate2_label_element.text,index_3,cate3_data_component_id,cate3_label_element.text,index_4,cate4_data_component_id,cate4_label_element.text])
                                                        print("****GET : ", index_4,"/",cate4_data_component_id,"/",cate4_label_element.text,"....completed")
                                                        index_4 += 1
                                                    #-------------------------------Get cate4---------------------------------------------------------------  
                                                    except (NoSuchElementException,TimeoutException,ElementClickInterceptedException, NoSuchAttributeException) as ec:
                                                        print(f"Selenium exception: {type(ec).__name__}")
                                                        ws.append([row[2].value, row[1].value, index, cate1_data_component_id, cate1_label_element.text, index_2, cate2_data_component_id, cate2_label_element.text,index_3,cate3_data_component_id,cate3_label_element.text,index_4,cate4_data_component_id,cate4_label_element.text])
                                                        print("--------End loop.Save%Go back to Cate 3")
                                                        break
                                                    except Exception as e:
                                                        # Raise a custom error if the element is not found
                                                        print(f"Sys error: {str(e)}")
                                                        print(f"Exception type: {sys.exc_info()[0]}")
                                                        break   
                                            #-------★click 열림    
                                            except (TimeoutException,ElementClickInterceptedException) as t:
                                                print(f"Selenium exception: {type(t).__name__}")
                                                print("....click unsuccessful.no tab exist. go to next siblings")
                                                pass
                                            except (NoSuchElementException, NoSuchAttributeException) as no:
                                                print(f"Selenium exception: {type(no).__name__}")
                                                print("--------End loop.Svae&Go back to Cate 2")
                                                ws.append([row[2].value, row[1].value, index, cate1_data_component_id, cate1_label_element.text, index_2, cate2_data_component_id, cate2_label_element.text,index_3,cate3_data_component_id,cate3_label_element.text,"","",""])
                                                break
                                            except Exception as e:
                                                # Raise a custom error if the element is not found
                                                print("--------Click cate3->cate4")
                                                print(f"Sys error: {str(e)}")
                                                print(f"Exception type: {sys.exc_info()[0]}")
                                                break
                                        #-------------------------------Get cate3---------------------------------------------------------------  
                                        except (NoSuchElementException,TimeoutException,ElementClickInterceptedException, NoSuchAttributeException) as ec:
                                            print(f"Selenium exception: {type(ec).__name__}")
                                            ws.append([row[2].value, row[1].value, index, cate1_data_component_id, cate1_label_element.text, index_2, cate2_data_component_id, cate2_label_element.text,index_3,cate3_data_component_id,cate3_label_element.text,"","",""])
                                            print("--------End loop.Save%Go back to Cate 2")
                                            break
                                        except Exception as e:
                                            # Raise a custom error if the element is not found
                                            print(f"Sys error: {str(e)}")
                                            print(f"Exception type: {sys.exc_info()[0]}")
                                            break                                                                       
                                #-------★click 열림    
                                except (TimeoutException,ElementClickInterceptedException) as t:
                                    print(f"Selenium exception: {type(t).__name__}")
                                    print("....click unsuccessful.no tab exist. go to next siblings")
                                    pass
                                except (NoSuchElementException, NoSuchAttributeException) as no:
                                    print(f"Selenium exception: {type(no).__name__}")
                                    print("--------End loop.Svae&Go back to Cate 1")
                                    ws.append([row[2].value, row[1].value, index, cate1_data_component_id, cate1_label_element.text, index_2, cate2_data_component_id, cate2_label_element.text,"", "", "","","",""])
                                    break
                                except Exception as e:
                                    # Raise a custom error if the element is not found
                                    print("--------Click cate2->cate3")
                                    print(f"Sys error: {str(e)}")
                                    print(f"Exception type: {sys.exc_info()[0]}")
                                    break
                            #-------------------------------Get cate2---------------------------------------------------------------  
                            except (NoSuchElementException, TimeoutException, ElementClickInterceptedException,NoSuchAttributeException) as ec:
                                print(f"Selenium exception: {type(ec).__name__}")
                                print("--------End loop.Svae&Go back to Cate1")
                                ws.append([row[2].value, row[1].value, index, cate1_data_component_id, cate1_label_element.text, index_2, cate2_data_component_id, cate2_label_element.text,"", "", "","","",""])
                                break
                            except Exception as e:
                                print("--------Get cate2")
                                print(f"Sys error: {str(e)}")
                                print(f"Exception type: {sys.exc_info()[0]}")
                                break
                    #-------★click 열림  
                    except (TimeoutException,ElementClickInterceptedException)  as t:
                        print(f"Selenium exception: {type(t).__name__}")
                        print("....click unsuccessful.no tab exist. go to next siblings")
                        pass
                    except (NoSuchElementException, NoSuchAttributeException) as no:
                        print(f"Selenium exception: {type(no).__name__}")
                        print("....click unsuccessful")
                        ws.append([row[2].value, row[1].value, index, cate1_data_component_id, cate1_label_element.text, "", "", "","","","","","",""])
                        break
                    except Exception as e:
                        print("--------Click cate1->cate2")
                        print(f"Sys error: {str(e)}")
                        print(f"Exception type: {sys.exc_info()[0]}")
                        print("No cate1 end")
                        break
                #-------------------------------Get cate1---------------------------------------------------------------   
                except (NoSuchElementException, TimeoutException,ElementClickInterceptedException, NoSuchAttributeException) as ec:
                    ws.append([row[2].value, row[1].value,"", "","", "", "", "","","","","","",""])
                    print(f"Selenium exception: {type(ec).__name__}")
                    print("--------End loop.Save & Go back to Mall")
                    break
                except Exception as e:
                    print("--------Get cate1")
                    print(f"Sys error: {str(e)}")
                    print(f"Exception type: {sys.exc_info()[0]}")
                    break
    #-------------------------------Inputput mall & donwload cate1-------------------------------------------------
        except (NoSuchElementException, TimeoutException,ElementClickInterceptedException, NoSuchAttributeException) as ec:
            print("--------cannot click the mall")
            print(f"Selenium exception: {type(ec).__name__}")
        except Exception as e:
            print("--------Get mall")
            print(f"Sys error: {str(e)}")
            print(f"Exception type: {sys.exc_info()[0]}")
            break
        wb.save("list_"+str(row[1].value)+"_"+row[0].value.replace("/","")+".xlsx")
        wb.close()
    else:
        print("pass mall cate....", row[2].value)
        pass